#!/usr/bin/env python3
"""
Script para importar dados de CLPs da planilha 'Planilha IPs.rev3.xlsx'
e gerar o arquivo clp_data.json para uso no IP Monitor.

Uso:
    python3 import_clp_data.py "/caminho/para/Planilha IPs.rev3.xlsx" [output.json]

Dependencias: openpyxl
"""

import sys
import re
import json
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERRO: openpyxl nao instalado. Execute: pip install openpyxl")
    sys.exit(1)


def normalizar_key(header):
    """Converte nome de coluna para key JSON padronizada."""
    mapa = {
        'Nº': 'numero',
        'Descrição': 'descricao',
        'Tipo': 'tipo',
        'Quadro Proteção': 'quadro_protecao',
        'Nº Circ.': 'num_circuito',
        'Quadro CLP': 'quadro_clp',
        'CLP Entrada': 'clp_entrada',
        'CLP Saída': 'clp_saida',
        'N° Relé': 'num_rele',
        'Módulo Expansão': 'modulo_expansao',
        'Porta I/O': 'porta_io',
    }
    return mapa.get(header, header.lower().replace(' ', '_').replace('/', '_').replace('°', ''))


def extrair_ip(texto):
    """Extrai endereco IP de uma string."""
    match = re.search(r'(\d+\.\d+\.\d+\.\d+)', texto)
    return match.group(1) if match else None


def extrair_modelo(texto):
    """Extrai modelo do CLP de uma string."""
    match = re.search(r'Micrologix\s+(\d+)\s+(\S+)', texto, re.IGNORECASE)
    if match:
        return f"Micrologix {match.group(1)} {match.group(2).rstrip('.')}"
    return "Micrologix (modelo desconhecido)"


def extrair_modulos(texto):
    """Extrai modulos de expansao da string."""
    return re.findall(r'(\d+x\d{4}-\w+)', texto)


def processar_aba(ws, sheet_name):
    """Processa uma aba da planilha e retorna dados do CLP."""
    # Linha 1 = titulo
    titulo = str(ws.cell(1, 1).value or '').strip()

    # Linha 2 = descricao
    descricao = str(ws.cell(2, 1).value or '').strip()

    # Linha 3 = modelo + IP
    linha3 = str(ws.cell(3, 1).value or '')
    ip = extrair_ip(linha3)
    if not ip:
        print(f"  AVISO: IP nao encontrado na aba '{sheet_name}', pulando...")
        return None

    modelo = extrair_modelo(linha3)
    modulos = extrair_modulos(linha3)

    # Linha 5 = cabecalhos
    headers_raw = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(5, col).value
        if val is not None:
            headers_raw.append(str(val).strip())
        else:
            headers_raw.append('')

    # Remove colunas vazias do final
    while headers_raw and not headers_raw[-1]:
        headers_raw.pop()

    # Tratar colunas duplicadas (ex: 'Nº' aparece 2x)
    # A segunda ocorrencia de 'Nº' refere-se ao numero do quadro/circuito
    headers_display = list(headers_raw)
    headers_keys = []
    seen = {}
    for h in headers_raw:
        key = normalizar_key(h)
        if key in seen:
            seen[key] += 1
            key = f"{key}_{seen[key]}"
        else:
            seen[key] = 1
        headers_keys.append(key)

    # Linhas 6+ = dados
    pontos = []
    for row in range(6, ws.max_row + 1):
        row_data = {}
        has_data = False

        for col_idx, key in enumerate(headers_keys):
            val = ws.cell(row, col_idx + 1).value
            if val is not None:
                has_data = True
                row_data[key] = str(val).strip()
            else:
                row_data[key] = ''

        if has_data and any(v for v in row_data.values()):
            pontos.append(row_data)

    return {
        'titulo': titulo,
        'descricao': descricao,
        'modelo': modelo,
        'ip': ip,
        'modulos_expansao': modulos,
        'aba_nome': sheet_name,
        'colunas_display': headers_display,
        'colunas_keys': headers_keys,
        'pontos': pontos,
        'total_pontos': len(pontos)
    }


def main():
    if len(sys.argv) < 2:
        print(f"Uso: {sys.argv[0]} <planilha.xlsx> [output.json]")
        sys.exit(1)

    planilha_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'clp_data.json'

    if not Path(planilha_path).exists():
        print(f"ERRO: Arquivo nao encontrado: {planilha_path}")
        sys.exit(1)

    print(f"Carregando planilha: {planilha_path}")
    wb = openpyxl.load_workbook(planilha_path, data_only=True)

    print(f"Total de abas: {len(wb.sheetnames)}")
    clp_data = {}

    for sheet_name in wb.sheetnames:
        if sheet_name == 'Resumo':
            print(f"  Pulando aba 'Resumo' (ja integrada)")
            continue

        print(f"  Processando: {sheet_name}...", end=' ')
        ws = wb[sheet_name]
        resultado = processar_aba(ws, sheet_name)

        if resultado is None:
            continue

        ip = resultado['ip']

        # Tratamento especial: IP compartilhado (ex: Reservatorio/Incendio)
        if ip in clp_data:
            print(f"IP duplicado ({ip}), mesclando com '{clp_data[ip]['aba_nome']}'")
            existente = clp_data[ip]

            # Adicionar secao separada para os pontos
            if 'secoes' not in existente:
                existente['secoes'] = [{
                    'nome': existente['aba_nome'],
                    'titulo': existente['titulo'],
                    'colunas_display': existente['colunas_display'],
                    'colunas_keys': existente['colunas_keys'],
                    'pontos': existente['pontos'],
                    'total_pontos': existente['total_pontos']
                }]
                existente['pontos'] = []
                existente['colunas_display'] = []
                existente['colunas_keys'] = []

            existente['secoes'].append({
                'nome': sheet_name,
                'titulo': resultado['titulo'],
                'colunas_display': resultado['colunas_display'],
                'colunas_keys': resultado['colunas_keys'],
                'pontos': resultado['pontos'],
                'total_pontos': resultado['total_pontos']
            })

            existente['titulo'] = f"{existente['secoes'][0]['nome']} / {sheet_name}"
            existente['total_pontos'] = sum(s['total_pontos'] for s in existente['secoes'])
        else:
            clp_data[ip] = resultado
            print(f"OK ({resultado['total_pontos']} pontos)")

    # Salvar JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(clp_data, f, indent=2, ensure_ascii=False)

    print(f"\nResultado:")
    print(f"  CLPs processados: {len(clp_data)}")
    total_pontos = sum(c['total_pontos'] for c in clp_data.values())
    print(f"  Total de pontos I/O: {total_pontos}")
    print(f"  Arquivo gerado: {output_path}")

    # Listar CLPs
    print(f"\nCLPs:")
    for ip in sorted(clp_data.keys(), key=lambda x: [int(p) for p in x.split('.')]):
        c = clp_data[ip]
        secoes = f" ({len(c['secoes'])} secoes)" if 'secoes' in c else ''
        print(f"  {ip:15s} | {c['titulo'][:50]:50s} | {c['total_pontos']:3d} pontos{secoes}")


if __name__ == '__main__':
    main()
